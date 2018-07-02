# -*- coding: utf-8 -*-
""" Project Jupyter - extract data from PDF by Vytautas"""

""" Importing libraries """
import pandas as pd
import os

""" Reading Dataset file """
os.chdir("C:\\Users\\Vytautas.Bielinskas\\Desktop\\Python\\")
DF = pd.read_csv("tabula-Statement (BULK) 02 July.csv", header=None)

""" <-------------  F U L L   D A T A   P R E P A R A T I O N -------------------->"""
def FullDataPreparation(DF):
    row = 0
    l = []
    while row < len(DF):
        
        for i in range(0, len(DF.columns), 1):
            if type(DF.iat[row, i]) != float:
                if DF.iat[row, i] == "Date":
                    IndexOfDate = i
                    date = True
                elif DF.iat[row, i] == "Description":
                    IndexOfDescription = i
                elif DF.iat[row, i] == "Property Address":
                    IndexOfAddress = i
                elif DF.iat[row,i] == "Paid out":
                    IndexOfPaidOut = i
                elif DF.iat[row,i] == "Paid in":
                    IndexOfPaidIn = i
                elif DF.iat[row,i] == "Balance":
                    IndexOfBalance = i
        
        if date == True:
            row = row +1
            while row+1 < len(DF) and DF.iat[row+1, IndexOfDate] != "Date":
                d = {}
                if type(DF.iat[row, IndexOfDate]) != float:
                    d["Date"] = DF.iat[row, IndexOfDate]
                    if DF.iat[row, IndexOfDescription][:12] == "Rent Receive":
                        d["Description"] = "Rent Received"
                    else:
                        if len(DF.iat[row, IndexOfDate]) > 3:
                            d["Description"] = DF.iat[row, IndexOfDescription]
                            if type(DF.iat[row+1, IndexOfDate]) == float:
                                takeInto = row + 1
                                while type(DF.iat[takeInto, IndexOfDate]) == float:
                                    d["Description"] = d["Description"] + " " + DF.iat[takeInto, IndexOfDescription] 
                                    takeInto = takeInto + 1
                        #d["Description"] = DF.iat[row, IndexOfDescription]
                    d["Property Address"] = DF.iat[row, IndexOfAddress]
                    d["Paid out"] = DF.iat[row, IndexOfPaidOut]
                    d["Paid in"] = DF.iat[row, IndexOfPaidIn]
                    d["Balance"] = DF.iat[row, IndexOfBalance]
                    d["Property"] = "-"
                    l.append(d)
                
                row = row + 1
        else:
            row = row + 1
        
        if row+1 == len(DF):
            d = {}
            d["Date"] = DF.iat[row, IndexOfDate]
            if DF.iat[row, IndexOfDescription][:12] == "Rent Receive":
                d["Description"] = "Rent Received" 
            else:
                d["Description"] = DF.iat[row, IndexOfDescription]
            d["Property Address"] = DF.iat[row, IndexOfAddress]
            d["Paid out"] = DF.iat[row, IndexOfPaidOut]
            d["Paid in"] = DF.iat[row, IndexOfPaidIn]
            d["Balance"] = DF.iat[row, IndexOfBalance]
            d["Property"] = "-"
            l.append(d)
        
    DFFilled = pd.DataFrame(l)
    DFFilled = DFFilled[["Date", "Description", "Property Address", "Paid out", "Paid in", "Balance", "Property"]]
    
    """ Set one Property name per block : start """
    DFFilled = DFFilled.rename(columns = {"Property Address" : "property"})
    namesT = DFFilled.property.unique() # 1st name is NaN
    names = []
    DFFilled = DFFilled.rename(columns = {"property" : "Property Address"})
    for i in range(1, len(namesT), 1):
        if namesT[i] != "Property Address":
            names.append(namesT[i])
    
    i = 0
    namesI = 0
    while i < len(DFFilled)-1:
        if DFFilled.iat[i, DFFilled.columns.get_loc("Description")] != "Description":
            if namesI < len(names):
                DFFilled.iat[i, DFFilled.columns.get_loc("Property")] = names[namesI]
                if i+1 < len(DFFilled)-1:
                    i = i +1
                else:
                    break
            else:
                break
        else:
            if i+1 < len(DFFilled)-1:
                namesI = namesI + 1
                i = i + 1
            else:
                break
    """ Set one Property name per block : end """
    
    return DFFilled
""" FULL DATA PREPARATION : END """

""" <---------------- B E F O R E   H A R D   W O R K ------------------------>"""
def PreparingDataForWork(DFFilled):
    import re
    
    """ Convert number-string values to Float"""
    for i in range(0, len(DFFilled), 1):#(0, len(DFFilled), 1)
        for j in range(2, len(DFFilled.columns), 1): # (2, len(DFFilled.columns), 1)
                if type(DFFilled.iat[i,j]) == str:
                    if "," in DFFilled.iat[i,j] or "." in DFFilled.iat[i,j]:
                        if "," in DFFilled.iat[i,j]:
                            DFFilled.iat[i,j] = DFFilled.iat[i,j].replace(",", ".")
                            if DFFilled.iat[i,j].count(".") == 2:
                                mid = round(len(DFFilled.iat[i,j])/2)
                                DFFilled.iat[i,j] = DFFilled.iat[i,j][:mid+1].replace(".", "") + DFFilled.iat[i,j][mid+1:]
                        DFFilled.iat[i,j] = float(DFFilled.iat[i,j])
                    if type(DFFilled.iat[i,j]) == str and re.search('[a-zA-Z]+', DFFilled.iat[i,j]) == None and DFFilled.iat[i,j] != "-":
                        DFFilled.iat[i,j] = float(DFFilled.iat[i,j])    
                                         
    """ Remove nan values from money fields """
    for i in range(0, len(DFFilled), 1):
        for j in range(2, len(DFFilled.columns), 1):
            if pd.isnull(DFFilled.iat[i,j]):
                DFFilled.iat[i,j] = float(0)
                
    """ Remove headers rows from the dataframe """
    rabbish = []
    for i in range(0, len(DFFilled), 1):
        if (DFFilled.iat[i, DFFilled.columns.get_loc("Description")]) == "Description":
            rabbish.append(i)
            
    DFFilled2 = DFFilled.drop(DFFilled.index[rabbish]).reset_index(drop=True)
            
    return DFFilled2
""" PREPARING DATA FOR WORK : END """

""" <---------------- F I L T E R   T H E   D A T A ------------------------>"""
""" FILTERING DATA : START """
def FilteringData(DFFilled):
    
    """ Removing used rows in DataFrame : start """
    def removeUsedRows(DataFrame):
        #print("Inside removeUsedRows")
        usedIndexes = []
        for i in range(0, len(DataFrame), 1):
            if DataFrame.iat[i, DFFilled.columns.get_loc("Property Address")] == "--data used--" or DataFrame.iat[i, DFFilled.columns.get_loc("Description")] == "Balance brought forward":
                usedIndexes.append(i)
    
        DataFrame2 = DataFrame.drop(DataFrame.index[usedIndexes]).reset_index(drop=True)
        return DataFrame2
    """ Removing used rows in DataFrame : end """
        
    AggDF01  = []
    AggDF02  = []
    AggDF03  = []
    AggDF04  = []
    AggDF05  = []
    AggDF06  = []
    AggDF07  = []
    AggDF08  = []
    AggDF09  = []
    AggDF10  = []
    AggDF11  = []
    AggDF12  = []
    AggDF13  = []
    AggDF13_ = []
    AggDF14  = []
    AggDF15  = []
    AggDF16  = []
    AggDF17  = []
    
    d01 = {} # Rent Received
    d02 = {} # Payment made to Owner
    d03 = {} # THTC Fee
    d04 = {} # Cleaning Inventory
    d05 = {} # Incomes from Landlord
    d06 = {} # Others (Paid out)
    d07 = {} # Agent Fee
    d08 = {} # Service Charge
    d09 = {} # Utility Bills (Water, gas, electricity)
    d10 = {} # Handyman
    d11 = {} # Other (Paid in)
    d12 = {} # Fund to diff Unit (Paid In)
    d13 = {} # Fund to diff Unit (Paid Out)
    d13_ = {} # Fund to diff Unit (to >>> : Paind in)
    d14 = {} # Insurance
    d15 = {} # Council Tax
    d16 = {} # Ground Rent (Paid Out)
    d17 = {} # Furniture & Fixing --> Beds
    
    for i in range(0, len(DFFilled)-1, 1):
        if "Rent Received".upper() in DFFilled["Description"][i].upper() or "Received from".upper() in DFFilled["Description"][i].upper() or "Balance Carried Forw".upper() in DFFilled["Description"][i]: #This place identify the payment (!!!)
            d01["Description"] = DFFilled["Description"][i]
            d01["Paid in"] = DFFilled["Paid in"][i]
            d01["Property Address"] = DFFilled["Property Address"][i]
            d01["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF01.append(dict(d01))
            
        elif "Payment made to Owner".upper() in DFFilled["Description"][i].upper(): # This place identify the payment (!!!)
            d02["Description"] = DFFilled["Description"][i]
            d02["Paid out"] = - DFFilled["Paid out"][i]
            d02["Property Address"] = DFFilled["Property"][i]
            d02["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF02.append(dict(d02))
            
        elif "Payment of Membe".upper() in DFFilled["Description"][i].upper() or "Membership Fee".upper() in DFFilled["Description"][i].upper(): # This place identify the payment (!!!)
            d03["Description"] = DFFilled["Description"][i]
            d03["Paid out"] = - DFFilled["Paid out"][i]
            d03["Property Address"] = DFFilled["Property"][i]
            d03["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF03.append(dict(d03))
            
        elif "Clean Eas".upper() in DFFilled["Description"][i].upper() or "Inventories".upper() in DFFilled["Description"][i].upper() or "need a clean".upper() in DFFilled["Description"][i].upper() or "clean and tidy".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d04["Description"] = DFFilled["Description"][i]
            d04["Paid out"] = - DFFilled["Paid out"][i]
            d04["Property Address"] = DFFilled["Property"][i]
            d04["Date"] = DFFilled["Date"][i]
            #print("Station 17")
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF04.append(dict(d04))
            
        elif "Deposit Protectio".upper() in DFFilled["Description"][i].upper() or "Happy Tenant".upper() in DFFilled["Description"][i].upper() or "Adam Lee Property Maintenance".upper() in DFFilled["Description"][i].upper() or "Chris Farragher".upper() in DFFilled["Description"][i].upper() or "Repair Aid".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d06["Description"] = DFFilled["Description"][i]
            d06["Paid out"] = - DFFilled["Paid out"][i]
            d06["Property Address"] = DFFilled["Property"][i]
            d06["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF06.append(dict(d06))
            
        elif "Coopers".upper() in DFFilled["Description"][i].upper() or "letting fee".upper() in DFFilled["Description"][i].upper() or "lettings".upper() in DFFilled["Description"][i].upper() or "Capital Lettings".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d07["Description"] = DFFilled["Description"][i]
            d07["Paid out"] = - DFFilled["Paid out"][i]
            d07["Property Address"] = DFFilled["Property"][i]
            d07["Date"] = DFFilled["Date"][i]
            
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF07.append(dict(d07))
            
        elif "Ringleys".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d08["Description"] = DFFilled["Description"][i]
            d08["Paid out"] = - DFFilled["Paid out"][i]
            d08["Property Address"] = DFFilled["Property"][i]
            d08["Date"] = DFFilled["Date"][i]
            #print("Station 32")
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF08.append(dict(d08))
            
        elif "EON (".upper() in DFFilled["Description"][i].upper() or "Affinity Water".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d09["Description"] = DFFilled["Description"][i]
            d09["Paid out"] = - DFFilled["Paid out"][i]
            d09["Property Address"] = DFFilled["Property"][i]
            d09["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF09.append(dict(d09))
            
        elif "Locksmith".upper() in DFFilled["Description"][i].upper() or "Maintenance call out".upper() in DFFilled["Description"][i].upper() or "Waste Concern".upper() in DFFilled["Description"][i].upper() or "property maintenance".upper() in DFFilled["Description"][i].upper() or "apple maintenance".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d10["Description"] = DFFilled["Description"][i]
            d10["Paid out"] = - DFFilled["Paid out"][i]
            d10["Property Address"] = DFFilled["Property"][i]
            d10["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF10.append(dict(d10))
        # continue
        elif "property maintenance".upper() in DFFilled["Description"][i].upper() and "ffix".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d10["Description"] = DFFilled["Description"][i]
            d10["Paid out"] = - DFFilled["Paid out"][i]
            d10["Property Address"] = DFFilled["Property"][i]
            d10["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF10.append(dict(d10))      
            
        elif "Dilapidations".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d11["Description"] = DFFilled["Description"][i]
            d11["Paid in"] = DFFilled["Paid in"][i]
            d11["Property Address"] = DFFilled["Property"][i]
            d11["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF11.append(dict(d11))
            #continue
        elif "key fobs".upper() in DFFilled["Description"][i].upper() and DFFilled["Paid in"][i] > 0: #This place identify the payment (!!!)
            d11["Description"] = DFFilled["Description"][i]
            d11["Paid in"] = DFFilled["Paid in"][i]
            d11["Property Address"] = DFFilled["Property"][i]
            d11["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF11.append(dict(d11))
            
        elif "Landlords Fun".upper() in DFFilled["Description"][i].upper(): #This place identify the payment (!!!)
            d11["Description"] = DFFilled["Description"][i]
            d11["Paid in"] = DFFilled["Paid in"][i]
            d11["Property Address"] = DFFilled["Property"][i]
            d11["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF11.append(dict(d11))
            
        elif "Transfer Funds For".upper() in DFFilled["Description"][i].upper() and DFFilled["Paid in"][i] > 0:
            d12["Description"] = DFFilled["Description"][i]
            d12["Paid in"] = DFFilled["Paid in"][i]
            d12["Property Address"] = DFFilled["Property"][i]
            d12["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF12.append(dict(d12))
            
        elif "Transfer from".upper() in DFFilled["Description"][i].upper() and DFFilled["Paid out"][i] > 0:
            
            property_from = DFFilled["Description"][i].split("from ")[1].split(" to")[0]
            property_to = DFFilled["Description"][i].split("to ")[1]
            try:
                property_to = property_to.split(" to ")[0]
            except:
                None
            
            print(DFFilled["Description"][i], ":::", property_from, ":::-->", property_to, ":: sum =", DFFilled["Paid out"][i] * (-1))
            print(DFFilled["Paid in"][i], "type: ", type(DFFilled["Paid out"][i]))
            d13["Description"] = DFFilled["Description"][i]
            d13["Paid out"] = float(DFFilled["Paid out"][i]) * (-1)
            d13["Property Address"] = property_from
            d13["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            # Dictionary d13_ for evaluating :::Property to:::
            d13_["Description"] = DFFilled["Description"][i]
            d13_["Paid in"] = DFFilled["Paid out"][i] 
            print("::", DFFilled["Paid out"][i])
            d13_["Property Address"] = property_to
            d13_["Date"] = DFFilled["Date"][i]
                        
            AggDF13.append(dict(d13))
            AggDF13_.append(dict(d13_))
                 
        elif "Lockton".upper() in DFFilled["Description"][i].upper() and "Insurance".upper() in DFFilled["Description"][i].upper():
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d14["Description"] = DFFilled["Description"][i]
            d14["Paid out"] = - DFFilled["Paid out"][i]
            d14["Property Address"] = DFFilled["Property"][i]
            d14["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF14.append(dict(d14))
            
        elif "coucil".upper() in DFFilled["Description"][i].upper() or "Hillingdon Council".upper() in DFFilled["Description"][i].upper():
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d15["Description"] = DFFilled["Description"][i]
            d15["Paid out"] = - DFFilled["Paid out"][i]
            d15["Property Address"] = DFFilled["Property"][i]
            d15["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF15.append(dict(d15))
            
        elif "transfer".upper() in DFFilled["Description"][i].upper():
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d15["Description"] = DFFilled["Description"][i]
            d15["Paid out"] = - DFFilled["Paid out"][i]
            d15["Property Address"] = DFFilled["Property"][i]
            d15["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF15.append(dict(d15))
            
        elif "Red Rock".upper() in DFFilled["Description"][i].upper() and DFFilled["Paid out"][i] == 150:
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d16["Description"] = DFFilled["Description"][i]
            d16["Paid out"] = - DFFilled["Paid out"][i]
            d16["Property Address"] = DFFilled["Property"][i]
            d16["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF16.append(dict(d16))
            
        elif "Red Rock".upper() in DFFilled["Description"][i].upper() and DFFilled["Paid out"][i] != 150:
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d08["Description"] = DFFilled["Description"][i]
            d08["Paid out"] = - DFFilled["Paid out"][i]
            d08["Property Address"] = DFFilled["Property"][i]
            d08["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF08.append(dict(d08))
            
        elif "We Fix Now".upper() in DFFilled["Description"][i].upper():
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d10["Description"] = DFFilled["Description"][i]
            d10["Paid out"] = - DFFilled["Paid out"][i]
            d10["Property Address"] = DFFilled["Property"][i]
            d10["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF10.append(dict(d10))
            
        elif "My Property Maintenance".upper() in DFFilled["Description"][i].upper():
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d10["Description"] = DFFilled["Description"][i]
            d10["Paid out"] = - DFFilled["Paid out"][i]
            d10["Property Address"] = DFFilled["Property"][i]
            d10["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF10.append(dict(d10))
            
        elif "Beds".upper() in DFFilled["Description"][i].upper():
            #print(DFFilled["Paid out"][i], "type: ", type(DFFilled["Paid out"][i]))
            d17["Description"] = DFFilled["Description"][i]
            d17["Paid out"] = - DFFilled["Paid out"][i]
            d17["Property Address"] = DFFilled["Property"][i]
            d17["Date"] = DFFilled["Date"][i]
            DFFilled["Property Address"][i] = "--data used--" # eliminating used row from the dataframe later
            
            AggDF17.append(dict(d17))
             
    #print("Station 47")
    try:
        DF01 = pd.DataFrame(AggDF01)
        DF01.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF01 = pd.DataFrame(columns = columns)
    
    try:
        DF02 = pd.DataFrame(AggDF02)
        DF02.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF02 = pd.DataFrame(columns = columns)
    
    try:
        DF03 = pd.DataFrame(AggDF03)
        DF03.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF03 = pd.DataFrame(columns = columns)
    
    try:
        DF04 = pd.DataFrame(AggDF04)
        DF04.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF04 = pd.DataFrame(columns = columns)
    
    try:
        DF05 = pd.DataFrame(AggDF05)
        DF05.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF05 = pd.DataFrame(columns = columns)
    
    try:
        DF06 = pd.DataFrame(AggDF06)
        DF06.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF06 = pd.DataFrame(columns = columns)
    
    try:
        DF07 = pd.DataFrame(AggDF07)
        DF07.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF07 = pd.DataFrame(columns = columns)        
    
    try:
        DF08 = pd.DataFrame(AggDF08)
        DF08.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF08 = pd.DataFrame(columns = columns)
        
    try:
        DF09 = pd.DataFrame(AggDF09)
        DF09.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF09 = pd.DataFrame(columns = columns)
    
    try:
        DF10 = pd.DataFrame(AggDF10)
        DF10.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF10 = pd.DataFrame(columns = columns)
    
    try:
        DF11 = pd.DataFrame(AggDF11)
        DF11.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
        #print("Station 60")
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF11 = pd.DataFrame(columns = columns) 
        
    try:
        DF12 = pd.DataFrame(AggDF12)
        DF12.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF12 = pd.DataFrame(columns = columns)
        
    try:
        DF13 = pd.DataFrame(AggDF13)
        DF13.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF13 = pd.DataFrame(columns = columns)
        
    try:
        DF13_ = pd.DataFrame(AggDF13_)
        DF13_.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF13_ = pd.DataFrame(columns = columns)
        
    try:
        DF14 = pd.DataFrame(AggDF14)
        DF14.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF14 = pd.DataFrame(columns = columns)
        
    try:
        DF15 = pd.DataFrame(AggDF15)
        DF15.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid in"] # if category is not exist in PDF
        DF15 = pd.DataFrame(columns = columns)
        
    try:
        DF16 = pd.DataFrame(AggDF16)
        DF16.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF16 = pd.DataFrame(columns = columns)
        
    try:
        DF17 = pd.DataFrame(AggDF17)
        DF16.sort_values(["Property Address", "Date"], ascending=[True, True], inplace=False)
    except:
        columns = ["Property Address", "Date", "Paid out"] # if category is not exist in PDF
        DF17 = pd.DataFrame(columns = columns)
    
    DFFilled = removeUsedRows(DFFilled)
    
    return DF01, DF02, DF03, DF04, DF05, DF06, DF07, DF08, DF09, DF10, DF11, DF12, DF13, DF13_, DF14, DF15, DF16, DF17, DFFilled
""" FILTERING DATA : END """

""" <---------------- A G R E G A T I O N   D A T A ------------------------> """
""" AGGREGATION <<Paid In>> : START """
def AggregationDataPaidIn(DF):
    if len(DF) > 0:
        DF.set_index(['Property Address', 'Date'])
        DF = DF.rename(columns = {"Paid in" : "paidin"})
        
        DF = DF.groupby(['Property Address', "Date"], as_index=False).paidin.sum()
        DF = DF.rename(columns = {"paidin" : "Paid in"})
    return DF
""" AGGREGATION <<Paid In>> : END """

""" AGGREGATION <<Paid Out>> : START """
def AggregationDataPaidOut(DF):
    if len(DF) > 0:
        DF.set_index(['Property Address', 'Date'])
        DF = DF.rename(columns = {"Paid out" : "paidout"})
        
        DF = DF.groupby(['Property Address', "Date"], as_index=False).paidout.sum()
        DF = DF.rename(columns = {"paidout" : "Paid out"})
    return DF
""" AGGREGATION <<Paid Out>> : END """

""" <---------------- S C E N A R I O ---------------------> """
""" #0 Scenario """
DFFilled = FullDataPreparation(DF)
DFFilled = PreparingDataForWork(DFFilled) #Preparing DataFrame with all the values

""" #1 Getting data : start """
PreparedDataByTypes = FilteringData(DFFilled)

DF_01  = PreparedDataByTypes[0] # Rent received & Received from
DF_02  = PreparedDataByTypes[1] # Payment to Owner
DF_03  = PreparedDataByTypes[2] # THTC Fee
DF_04  = PreparedDataByTypes[3] # Cleaning and Inventory Check In&Out
DF_05  = PreparedDataByTypes[4] # Incomes from Landlord
DF_06  = PreparedDataByTypes[5] # Others (Paid out - EXPENSES)
DF_07  = PreparedDataByTypes[6] # Agent Fee
DF_08  = PreparedDataByTypes[7] # Service Charge
DF_09  = PreparedDataByTypes[8] # Utility Bills (Water, gas, electricity)
DF_10  = PreparedDataByTypes[9] # Handyman
DF_11  = PreparedDataByTypes[10] # Others (Paid in - INCOMES)
DF_12  = PreparedDataByTypes[11] # Fund to diff Unit (Paid In)
DF_13  = PreparedDataByTypes[12] # Fund to diff Unit (Paid Out)
DF_13_ = PreparedDataByTypes[13] # Fund to diff Unit (Paid In) : updated at 2018 05 03
DF_14  = PreparedDataByTypes[14] # Insurance
DF_15  = PreparedDataByTypes[15] # Council Tax
DF_16  = PreparedDataByTypes[16] # Ground Rent
DF_17  = PreparedDataByTypes[17] # Beds

DF_01  = AggregationDataPaidIn(DF_01)    # Rent reveiced & Received from
DF_02  = AggregationDataPaidOut(DF_02)   # Payment to Owner
DF_03  = AggregationDataPaidOut(DF_03)   # THTC Fee
DF_04  = AggregationDataPaidOut(DF_04)   # Cleaning and Inventory Check In&Out
DF_05  = AggregationDataPaidIn(DF_05)    # Incomes from Landlord
DF_06  = AggregationDataPaidOut(DF_06)   # Others (Paid out - EXPENSES)
DF_07  = AggregationDataPaidOut(DF_07)   # Agent Fee
DF_08  = AggregationDataPaidOut(DF_08)   # Service Charge
DF_09  = AggregationDataPaidOut(DF_09)   # Utility Bills (Water, gas, electricity)
DF_10  = AggregationDataPaidOut(DF_10)   # Handyman
DF_11  = AggregationDataPaidIn(DF_11)    # Others (Paid in - INCOMES)
DF_12  = AggregationDataPaidIn(DF_12)    # Fund to diff Unit (Paid In)
DF_13  = AggregationDataPaidOut(DF_13)   # Fund to diff Unit (Paid Out)
DF_13_ = AggregationDataPaidIn(DF_13_)   # Fund to diff Unit (Paid In) : updated at 2018 05 03
DF_14  = AggregationDataPaidOut(DF_14)   # Insurance
DF_15  = AggregationDataPaidOut(DF_15)   # Insurance
DF_16  = AggregationDataPaidOut(DF_16)   # Ground Rent
DF_17  = AggregationDataPaidOut(DF_17)   # Beds

DF01_RentReceivedFromTenant = pd.DataFrame(DF_01)
DF02_PaymentMadeToOwner = pd.DataFrame(DF_02)
DF03_THTCFee = pd.DataFrame(DF_03)
DF04_CleaningInventory = pd.DataFrame(DF_04)
DF05_IncomesFromLandlord = pd.DataFrame(DF_05)
DF06_OthersExpenses = pd.DataFrame(DF_06)
DF07_AgentFee = pd.DataFrame(DF_07)
DF08_ServiceCharge = pd.DataFrame(DF_08)
DF09_UtilityBills = pd.DataFrame(DF_09)
DF10_Handyman = pd.DataFrame(DF_10)
DF11_OtherIncomes = pd.DataFrame(DF_11)
DF12_FundToDiffUnit = pd.DataFrame(DF_12)
DF13_FundToDiffUnitOut = pd.DataFrame(DF_13)
DF13__FundToDiffUnitIn = pd.DataFrame(DF_13_)
DF14_Insurance = pd.DataFrame(DF_14)
DF15_CouncilTax = pd.DataFrame(DF_15)
DF16_GroundRent = pd.DataFrame(DF_16)
DF17_Beds = pd.DataFrame(DF_17)
""" #1 Getting data : end """

""" <--------------- W R I T E   D A T A  T O   E X C E L ---------------------> """
""" Write Data to Excel : START """
def ExcelPart(DF01_RentReceivedFromTenant, DF02_PaymentMadeToOwner, DF03_THTCFee, DF04_CleaningInventory, DF05_IncomesFromLandlord, DF06_OthersExpenses, DF07_AgentFee, DF08_ServiceCharge, DF09_UtilityBills, DF10_Handyman, DF11_OtherIncomes, DF12_FundToDiffUnit, DF13_FundToDiffUnitOut, DF13__FundToDiffUnitIn, DF14_Insurance, DF15_CouncilTax, DF16_GroundRent, DF17_Beds):
    import openpyxl, os, datetime
    from pandas import ExcelWriter as ewriter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
    
    """ Hard part in Excel : start"""
    def writeDataExcel(startRow, propertyName, Data, sheet, CategoryRowIndex):
        
        print("writing....")
        if propertyName == "CommonPartsOtter":
            print("CommonPartsOtter rasta")
        startMonthCellIndex = 1039 #This cell means ACL [766] 2017 09 01, ADP [796] 2017 10 01, AFY [857] 2017 12 01, [888]  2018 01 01, [947] 2018 03 01, [978] 2018 04 01, [1008] 2018 05 01, [1039] 2018 06 01
        
        startMonthDay = float(str(sheet.cell(row = 9, column = startMonthCellIndex).value).split(" ")[0][-2:])
        #print("startRow = ", startRow, ", propertyName = ", propertyName, "RowID = ", Data.loc[Data["Property Address"] == propertyName].index.values, "len = ", len(Data.loc[Data["Property Address"] == propertyName].index.values), "Month starts: ", startMonthDay)    
        
        # Give color to marked cells
        fillGreen = PatternFill(fill_type = "solid",
                                start_color = "009e73",
                                end_color = "009e73")
        
        fillYellow = PatternFill(fill_type = "solid",
                                start_color = "d9f442",
                                end_color = "d9f442")
        
        if Data.iat[Data.loc[Data["Property Address"] == propertyName].index.values[0], 2] > 0:
            print(propertyName, "---> Paid In")
            for i in range(0, len(Data.loc[Data["Property Address"] == propertyName].index.values), 1):
                index = Data.loc[Data["Property Address"] == propertyName].index.values[i]
                #print("Date :", Data.iat[index, 1], "Date day: ", float(Data.iat[index, 1][:2]), "Paid in :", Data.iat[index, 2])
                
                print("Error: --->", propertyName, CategoryRowIndex, (Data.iat[index, 1][:2]), " : ", Data.iat[index, 2])
                #print(Data)
                sheet.cell(row = startRow + CategoryRowIndex, column = startMonthCellIndex + float(str(Data.iat[index, 1][:2]).replace("/", "")) - 1).value = Data.iat[index, 2]
                sheet.cell(row = startRow + CategoryRowIndex, column = startMonthCellIndex + float(str(Data.iat[index, 1][:2]).replace("/", "")) - 1).fill = fillGreen
        else:
            print(propertyName, "---> Paid Out")        
            print("DATA :-->", Data)
            print("Lenght:", len(Data))
            for i in range(0, len(Data.loc[Data["Property Address"] == propertyName].index.values), 1):
                index = Data.loc[Data["Property Address"] == propertyName].index.values[i]                                             # Get number of row that correspond the Property Name
                print("Data.iat[index, 2] =", Data.iat[index, 2])
                print("Maybe error is here:", str(Data.iat[index, 1][:2]).replace("/", ""))
                #print("Date :", Data.iat[index, 1], "Date day: ", float(Data.iat[index, 1][:2]), "Paid out :", Data.iat[index, 2])
                    
                sheet.cell(row = startRow + CategoryRowIndex, column = startMonthCellIndex + float(str(Data.iat[index, 1][:2]).replace("/", "")) - 1).value = Data.iat[index, 2]
                sheet.cell(row = startRow + CategoryRowIndex, column = startMonthCellIndex + float(str(Data.iat[index, 1][:2]).replace("/", "")) - 1).fill = fillYellow
        
        for i in range(0, len(Data.loc[Data["Property Address"] == propertyName].index.values), 1):
            index = Data.loc[Data["Property Address"] == propertyName].index.values[i]
        
        print("")
        
        return sheet
    """ Hard part in Excel : end"""
    
    """Prepare primary Excel File"""
    os.chdir("C:\\Users\\Vytautas.Bielinskas\\Desktop\\Python\\02 Jupyter\\")
    
    filename = "primary_file.xlsx"
    wb = openpyxl.load_workbook(filename, data_only = True)
    
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])   # If file is original --> 4
    
    Properties = ["16 Otter Way","17 Otter Way", "19 Otter Way", "20 Otter Way", "22 Otter Way", "23 Otter Way", "24 Otter Way",
              "26 Otter Way", "29 Otter Way", "30 Otter Way", "31 Otter Way", "36 Otter Way", "37 Otter Way", 
             "41 Otter Way", "42 Otter Way", "45 Otter Way", "46 Otter Way", "47 Otter Way", "48 Otter Way", "50 Otter Way", 
             "26 Autumn Way", "42 Autumn Way", "56 Autumn Way", "62 Autumn Way", "Common Parts Otter Way", 
              "Plot 1, Frays Court, Swan Road", "Plot 2, Frays Court, Swan Road", "Plot 3, Frays Court, Swan Road", 
             "Plot 4, Frays Court, Swan Road", "Plot 5, Frays Court, Swan Road", "Plot 6, Frays Court, Swan Road", 
             "Plot 7, Frays Court, Swan Road", "Plot 8, Frays Court, Swan Road", "Plot 9, Frays Court, Swan Road",
             "Plot 10, Frays Court, Swan Road", "Plot 11, Frays Court, Swan Road", "Plot 12, Frays Court, Swan Road",
             "Plot 13, Frays Court, Swan Road", "Plot 14, Frays Court, Swan Road", "Plot 15, Frays Court, Swan Road",
             "Common Parts, Frays Court, Swan Road"]
    
    PropertiesIndexes = {}
    ListOfProperties = []
    
    for i in range(1, sheet.max_row, 1):
        for j in range(0, len(Properties), 1):
            if Properties[j] == sheet.cell(row = i, column = 1).value:
                PropertiesIndexes["Property"] = Properties[j]
                PropertiesIndexes["Row in Excel"] = i
                ListOfProperties.append(dict(PropertiesIndexes))
                
    DFProperties = pd.DataFrame(ListOfProperties)
    DFProperties = DFProperties.drop_duplicates(subset=None, keep='first', inplace=False)

    # Make the names of Properties more similar!
    for i in range(0, len(DFProperties), 1):
        print(i, DFProperties.iat[i,0])
        if "swan".upper() in DFProperties.iat[i,0].upper():
            DFProperties.iat[i,0] = DFProperties.iat[i,0].split("Swan")[0]
        if "plot".upper() in DFProperties.iat[i,0].upper():
            DFProperties.iat[i,0] = DFProperties.iat[i,0].split("Plot")[1]
        DFProperties.iat[i,0] = DFProperties.iat[i,0].replace(", ", " ").split(" Court")[0]
        DFProperties.iat[i,0] = DFProperties.iat[i,0].replace(" ", "")
        if "Common Parts Otter Way".upper() in DFProperties.iat[i,0].upper():
            DFProperties.iat[i,0] = "CommonPartsOtter" # Common Parts Otter
            print("COMMON PARTS OTTER:",  DFProperties.iat[i,0])

    # Making shorter names for Frays properties in PDF dataset (DF-01)
    for i in range(0, len(DF01_RentReceivedFromTenant), 1):
        if "court".upper() in DF01_RentReceivedFromTenant.iat[i,0].upper():
            DF01_RentReceivedFromTenant.iat[i,0] = DF01_RentReceivedFromTenant.iat[i,0].split(" Court")[0]
        DF01_RentReceivedFromTenant.iat[i,0] = DF01_RentReceivedFromTenant.iat[i,0].replace(" ", "")
        #print(i, " : ", DF01_RentReceivedFromTenant.iat[i,0]," ", type(DF01_RentReceivedFromTenant.iat[i,0]))
        
    # Making shorter names for Frays properties in PDF dataset (DF-02)
    for i in range(0, len(DF02_PaymentMadeToOwner), 1):
        #print(i, " : ", DF02_PaymentMadeToOwner.iat[i,0]," ", type(DF02_PaymentMadeToOwner.iat[i,0]))
        if "court".upper() in DF02_PaymentMadeToOwner.iat[i,0].upper():
            DF02_PaymentMadeToOwner.iat[i,0] = DF02_PaymentMadeToOwner.iat[i,0].split(" Court")[0]
        DF02_PaymentMadeToOwner.iat[i,0] = DF02_PaymentMadeToOwner.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-03)
    for i in range(0, len(DF03_THTCFee), 1):
        #print(i, " : ", DF03_THTCFee.iat[i,0]," ", type(DF03_THTCFee.iat[i,0]))
        if "court".upper() in DF03_THTCFee.iat[i,0].upper():
            DF03_THTCFee.iat[i,0] = DF03_THTCFee.iat[i,0].split(" Court")[0]
        DF03_THTCFee.iat[i,0] = DF03_THTCFee.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-04)
    for i in range(0, len(DF04_CleaningInventory), 1):
        #print(i, " : ", DF04_CleaningInventory.iat[i,0]," ", type(DF04_CleaningInventory.iat[i,0]))
        if "court".upper() in DF04_CleaningInventory.iat[i,0].upper():
            DF04_CleaningInventory.iat[i,0] = DF04_CleaningInventory.iat[i,0].split(" Court")[0]
        DF04_CleaningInventory.iat[i,0] = DF04_CleaningInventory.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-05)
    for i in range(0, len(DF05_IncomesFromLandlord), 1):
        #print(i, " : ", DF05_IncomesFromLandlord.iat[i,0]," ", type(DF05_IncomesFromLandlord.iat[i,0]))
        if "court".upper() in DF05_IncomesFromLandlord.iat[i,0].upper():
            DF05_IncomesFromLandlord.iat[i,0] = DF05_IncomesFromLandlord.iat[i,0].split(" Court")[0]
        DF05_IncomesFromLandlord.iat[i,0] = DF05_IncomesFromLandlord.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-06)
    for i in range(0, len(DF06_OthersExpenses), 1):
        #print(i, " : ", DF06_OthersExpenses.iat[i,0]," ", type(DF06_OthersExpenses.iat[i,0]))
        if "court".upper() in DF06_OthersExpenses.iat[i,0].upper():
            DF06_OthersExpenses.iat[i,0] = DF06_OthersExpenses.iat[i,0].split(" Court")[0]
        DF06_OthersExpenses.iat[i,0] = DF06_OthersExpenses.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-07)
    for i in range(0, len(DF07_AgentFee), 1):
        #print(i, " : ", DF07_AgentFee.iat[i,0]," ", type(DF07_AgentFee.iat[i,0]))
        if "court".upper() in DF07_AgentFee.iat[i,0].upper():
            DF07_AgentFee.iat[i,0] = DF07_AgentFee.iat[i,0].split(" Court")[0]
        DF07_AgentFee.iat[i,0] = DF07_AgentFee.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-08)
    for i in range(0, len(DF08_ServiceCharge), 1):
        #print(i, " : ", DF08_ServiceCharge.iat[i,0]," ", type(DF08_ServiceCharge.iat[i,0]))
        if "court".upper() in DF08_ServiceCharge.iat[i,0].upper():
            DF08_ServiceCharge.iat[i,0] = DF08_ServiceCharge.iat[i,0].split(" Court")[0]
        DF08_ServiceCharge.iat[i,0] = DF08_ServiceCharge.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-09)
    for i in range(0, len(DF09_UtilityBills), 1):
        #print(i, " : ", DF09_UtilityBills.iat[i,0]," ", type(DF09_UtilityBills.iat[i,0]))
        if "court".upper() in DF09_UtilityBills.iat[i,0].upper():
            DF09_UtilityBills.iat[i,0] = DF09_UtilityBills.iat[i,0].split(" Court")[0]
        DF09_UtilityBills.iat[i,0] = DF09_UtilityBills.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-10)
    for i in range(0, len(DF10_Handyman), 1):
        print(i, " : ", DF10_Handyman.iat[i,0]," ", type(DF10_Handyman.iat[i,0]))
        if "court".upper() in DF10_Handyman.iat[i,0].upper():
            DF10_Handyman.iat[i,0] = DF10_Handyman.iat[i,0].split(" Court")[0]
        DF10_Handyman.iat[i,0] = DF10_Handyman.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-11)
    for i in range(0, len(DF11_OtherIncomes), 1):
        #print(i, " : ", DF11_OtherIncomes.iat[i,0]," ", type(DF11_OtherIncomes.iat[i,0]))
        if "court".upper() in DF11_OtherIncomes.iat[i,0].upper():
            DF11_OtherIncomes.iat[i,0] = DF11_OtherIncomes.iat[i,0].split(" Court")[0]
        DF11_OtherIncomes.iat[i,0] = DF11_OtherIncomes.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-12)
    for i in range(0, len(DF12_FundToDiffUnit), 1):
        #print(i, " : ", DF12_FundToDiffUnit.iat[i,0]," ", type(DF12_FundToDiffUnit.iat[i,0]))
        if "court".upper() in DF12_FundToDiffUnit.iat[i,0].upper():
            DF12_FundToDiffUnit.iat[i,0] = DF12_FundToDiffUnit.iat[i,0].split(" Court")[0]
        DF12_FundToDiffUnit.iat[i,0] = DF12_FundToDiffUnit.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-13)
    for i in range(0, len(DF13_FundToDiffUnitOut), 1):
        #print(i, " : ", DF13_FundToDiffUnitOut.iat[i,0]," ", type(DF13_FundToDiffUnitOut.iat[i,0]))
        if "court".upper() in DF13_FundToDiffUnitOut.iat[i,0].upper():
            DF13_FundToDiffUnitOut.iat[i,0] = DF13_FundToDiffUnitOut.iat[i,0].split(" Court")[0]
        DF13_FundToDiffUnitOut.iat[i,0] = DF13_FundToDiffUnitOut.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-_13)
    for i in range(0, len(DF13__FundToDiffUnitIn), 1):
        #print(i, " : ", DF13__FundToDiffUnitIn.iat[i,0]," ", type(DF13__FundToDiffUnitIn.iat[i,0]))
        if "court".upper() in DF13__FundToDiffUnitIn.iat[i,0].upper():
            DF13__FundToDiffUnitIn.iat[i,0] = DF13__FundToDiffUnitIn.iat[i,0].split(" Court")[0]
        DF13__FundToDiffUnitIn.iat[i,0] = DF13__FundToDiffUnitIn.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-14)
    for i in range(0, len(DF14_Insurance), 1):
        #print(i, " : ", DF14_Insurance.iat[i,0]," ", type(DF14_Insurance.iat[i,0]))
        if "court".upper() in DF14_Insurance.iat[i,0].upper():
            DF14_Insurance.iat[i,0] = DF14_Insurance.iat[i,0].split(" Court")[0]
        DF14_Insurance.iat[i,0] = DF14_Insurance.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-15)
    for i in range(0, len(DF15_CouncilTax), 1):
        #print(i, " : ", DF15_CouncilTax.iat[i,0]," ", type(DF15_CouncilTax.iat[i,0]))
        if "court".upper() in DF15_CouncilTax.iat[i,0].upper():
            DF15_CouncilTax.iat[i,0] = DF15_CouncilTax.iat[i,0].split(" Court")[0]
        DF15_CouncilTax.iat[i,0] = DF15_CouncilTax.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-16)
    for i in range(0, len(DF16_GroundRent), 1):
        #print(i, " : ", DF16_GroundRent.iat[i,0]," ", type(DF16_GroundRent.iat[i,0]))
        if "court".upper() in DF16_GroundRent.iat[i,0].upper():
            DF16_GroundRent.iat[i,0] = DF16_GroundRent.iat[i,0].split(" Court")[0]
        DF16_GroundRent.iat[i,0] = DF16_GroundRent.iat[i,0].replace(" ", "")
        
    # Making shorter names for Frays properties in PDF dataset (DF-17)
    for i in range(0, len(DF17_Beds), 1):
        #print(i, " : ", DF17_GroundRent.iat[i,0]," ", type(DF17_GroundRent.iat[i,0]))
        if "court".upper() in DF17_Beds.iat[i,0].upper():
            DF17_Beds.iat[i,0] = DF17_Beds.iat[i,0].split(" Court")[0]
        DF17_Beds.iat[i,0] = DF17_Beds.iat[i,0].replace(" ", "")

    print(" ")
    # <<<--- W R I T I N G   D A T A   T O   E X C E L --->>>
    CategoryRowIndex = 8 # StartRow + 8 in Excel for Rent Received From Tenant
    for i in range(0, len(DFProperties), 1):
        for j in range(0, len(DF01_RentReceivedFromTenant), 1):
            if DFProperties.iat[i,0] == DF01_RentReceivedFromTenant.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF01_RentReceivedFromTenant, sheet, CategoryRowIndex)
    
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 19 # StartRow + 19 in Excel for Payment Made To Owner
        for j in range(0, len(DF02_PaymentMadeToOwner), 1):
            if DFProperties.iat[i,0] == DF02_PaymentMadeToOwner.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF02_PaymentMadeToOwner, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 24 # StartRow + 24 in Excel for THTC Fee
        for j in range(0, len(DF03_THTCFee), 1):
            if DFProperties.iat[i,0] == DF03_THTCFee.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF03_THTCFee, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 26 # StartRow + 26 in Excel for Cleaning Inventory
        for j in range(0, len(DF04_CleaningInventory), 1):
            if DFProperties.iat[i,0] == DF04_CleaningInventory.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF04_CleaningInventory, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 4 # StartRow + 4 in Excel for Incomes From Landlord
        for j in range(0, len(DF05_IncomesFromLandlord), 1):
            if DFProperties.iat[i,0] == DF05_IncomesFromLandlord.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF05_IncomesFromLandlord, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 28 # StartRow + 28 in Excel for Other Expenses
        for j in range(0, len(DF06_OthersExpenses), 1):
            if DFProperties.iat[i,0] == DF06_OthersExpenses.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF06_OthersExpenses, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 23 # StartRow + 23 in Excel for Agent Fee
        for j in range(0, len(DF07_AgentFee), 1):
            if DFProperties.iat[i,0] == DF07_AgentFee.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF07_AgentFee, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 21 # StartRow + 23 in Excel for Service Charge
        for j in range(0, len(DF08_ServiceCharge), 1):
            if DFProperties.iat[i,0] == DF08_ServiceCharge.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF08_ServiceCharge, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 22 # StartRow + 22 in Excel for Utility Bills
        for j in range(0, len(DF09_UtilityBills), 1):
            if DFProperties.iat[i,0] == DF09_UtilityBills.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF09_UtilityBills, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 25 # StartRow + 25 in Excel for Handyman
        for j in range(0, len(DF10_Handyman), 1):
            if DFProperties.iat[i,0] == DF10_Handyman.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF10_Handyman, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 11 # StartRow + 11 in Excel for Other Incomes
        for j in range(0, len(DF11_OtherIncomes), 1):
            if DFProperties.iat[i,0] == DF11_OtherIncomes.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF11_OtherIncomes, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 6 # StartRow + 6 in Excel for Funds from Rental Income of diff Unit
        for j in range(0, len(DF12_FundToDiffUnit), 1):
            if DFProperties.iat[i,0] == DF12_FundToDiffUnit.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF12_FundToDiffUnit, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 15 # StartRow + 15 in Excel for Funds from Rental Income of diff Unit (Expenditures)
        for j in range(0, len(DF13_FundToDiffUnitOut), 1):
            if DFProperties.iat[i,0] == DF13_FundToDiffUnitOut.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF13_FundToDiffUnitOut, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 6 # StartRow + 6 in Excel for Funds from Rental Income of diff Unit (Expenditures)
        for j in range(0, len(DF13__FundToDiffUnitIn), 1):
            if DFProperties.iat[i,0] == DF13__FundToDiffUnitIn.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF13__FundToDiffUnitIn, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 29 # StartRow + 29 in Excel for Insurance
        for j in range(0, len(DF14_Insurance), 1):
            if DFProperties.iat[i,0] == DF14_Insurance.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF14_Insurance, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 27 # StartRow + 27 in Excel for Council Tax
        for j in range(0, len(DF15_CouncilTax), 1):
            if DFProperties.iat[i,0] == DF15_CouncilTax.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF15_CouncilTax, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 20 # StartRow + 38 in Excel for Council Tax
        for j in range(0, len(DF16_GroundRent), 1):
            if DFProperties.iat[i,0] == DF16_GroundRent.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF16_GroundRent, sheet, CategoryRowIndex)
                
    for i in range(0, len(DFProperties), 1):
        CategoryRowIndex = 31 # StartRow + 31 in Excel for Council Tax
        for j in range(0, len(DF17_Beds), 1):
            if DFProperties.iat[i,0] == DF17_Beds.iat[j,0]:
                # Here is starting writing Data to Excel             
                writeDataExcel(DFProperties.iat[i,1], DFProperties.iat[i,0], DF17_Beds, sheet, CategoryRowIndex)
                
    """ Write formulas to the file : start """
    asset = {}
    properties = []
    id_object = 0
    
    print("Writing formulas now!")
    for rows in range (17, sheet.max_row+1, 1):
        value = sheet.cell(row = rows, column = 1).value
        if value == "Bop":
            bop = rows
            asset["Bop"] = bop
            
            total_income = bop + 11
            asset["Total Income"] = total_income
            
            total_expenditure = total_income + 41
            asset["Total Expenditure"] = total_expenditure
            
            eop = total_expenditure + 1
            asset["Eop"] = eop 
            
            for column in range(698, sheet.max_column+1, 1):
                formula_bop = '=' + openpyxl.utils.get_column_letter(column-1) + str(eop)
                sheet.cell(row = bop, column = column).value = formula_bop
                
                formula_total_income = '=SUM(' + openpyxl.utils.get_column_letter(column) + str(total_income - 8) + ':' + openpyxl.utils.get_column_letter(column) + str(total_income-1) + ')'
                sheet.cell(row = total_income, column = column).value = formula_total_income
                
                formula_total_exp = '=SUM(' + openpyxl.utils.get_column_letter(column) + str(total_income + 1) + ':' + openpyxl.utils.get_column_letter(column) + str(total_expenditure-1) + ')'
                sheet.cell(row = total_expenditure, column = column).value = formula_total_exp
                
                formula_eop = '=SUM(' + openpyxl.utils.get_column_letter(column) + str(total_income) + "," + openpyxl.utils.get_column_letter(column) + str(bop) + "," + openpyxl.utils.get_column_letter(column) + str(total_expenditure) + ')'
                print(formula_eop)
                sheet.cell(row = eop, column = column).value = formula_eop
                
            formula_total_income_sum = '=SUM(' + 'C'+ str(total_income) + ':' + str(openpyxl.utils.get_column_letter(sheet.max_column)) + str(total_income) + ')'
            sheet.cell(row = total_income, column = 2).value = formula_total_income_sum
            
            formula_total_expend_sum = '=SUM(' + 'C'+ str(total_expenditure) + ':' + str(openpyxl.utils.get_column_letter(sheet.max_column)) + str(total_expenditure) + ')'
            sheet.cell(row = total_expenditure, column = 2).value = formula_total_expend_sum
            
            properties.append(dict(asset))
            print("Bop", bop, ", Total Income", total_income, ", Total Expenditure", total_expenditure, ", Eop", eop)    
    """ Write formulas to the file : end """
    
    wb.save("RESULT_FILE_final_formules_2_vsio.xlsx")
    
    return None
""" Write Data to Excel : END """

ItemsLeft = PreparedDataByTypes[18] # Take the last parameter to get cleaned DataFrame
if len(ItemsLeft) > 0:
    print("Prepare fro EXCEL part >>>")
    ExcelPart(DF01_RentReceivedFromTenant, 
              DF02_PaymentMadeToOwner, 
              DF03_THTCFee, 
              DF04_CleaningInventory,
              DF05_IncomesFromLandlord, 
              DF06_OthersExpenses,
              DF07_AgentFee, 
              DF08_ServiceCharge,
              DF09_UtilityBills, 
              DF10_Handyman, 
              DF11_OtherIncomes,
              DF12_FundToDiffUnit,
              DF13_FundToDiffUnitOut,
              DF13__FundToDiffUnitIn,
              DF14_Insurance,
              DF15_CouncilTax,
              DF16_GroundRent,
              DF17_Beds)
else:
    print("Check the list!")